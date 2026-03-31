from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from database import get_db
from models import Transaction, Category
import io

router = APIRouter(prefix="/api/export", tags=["export"])

@router.get("/excel/{year}")
def export_year_excel(year: int, db: Session = Depends(get_db)):
    """Export all transactions for a year as Excel file."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        # Fallback to CSV if openpyxl not available
        return export_year_csv(year, db)

    wb = openpyxl.Workbook()

    categories = db.query(Category).order_by(Category.display_order).all()
    cat_map = {c.id: c for c in categories}

    MESI = ["Gennaio", "Febbraio", "Marzo", "Aprile", "Maggio", "Giugno",
            "Luglio", "Agosto", "Settembre", "Ottobre", "Novembre", "Dicembre"]

    # Sheet 1: Riepilogo annuale (categories as rows, months as columns)
    ws = wb.active
    ws.title = f"Riepilogo {year}"

    # Header row
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")

    ws.cell(row=1, column=1, value="Categoria").font = header_font
    ws.cell(row=1, column=1).fill = header_fill
    ws.column_dimensions['A'].width = 22

    for i, mese in enumerate(MESI):
        cell = ws.cell(row=1, column=i+2, value=mese)
        cell.font = header_font
        cell.fill = header_fill
        ws.column_dimensions[chr(66+i) if i < 25 else 'AA'].width = 14

    total_col = 14
    cell = ws.cell(row=1, column=total_col, value="TOTALE")
    cell.font = header_font
    cell.fill = header_fill

    # Data rows - one per category
    for row_idx, cat in enumerate(categories, start=2):
        ws.cell(row=row_idx, column=1, value=cat.name)
        row_total = 0
        for m in range(1, 13):
            txs = db.query(Transaction).filter(
                Transaction.year == year,
                Transaction.month == m,
                Transaction.category_id == cat.id
            ).all()
            total = sum(t.amount for t in txs)
            if total != 0:
                ws.cell(row=row_idx, column=m+1, value=round(total, 2))
            row_total += total
        if row_total != 0:
            ws.cell(row=row_idx, column=total_col, value=round(row_total, 2))

    # Totals row
    total_row = len(categories) + 2
    ws.cell(row=total_row, column=1, value="TOTALE").font = Font(bold=True)

    # Sheet 2: Dettaglio transazioni
    ws2 = wb.create_sheet("Dettaglio")
    headers2 = ["Mese", "Categoria", "Tipo", "Descrizione", "Importo", "Fonte"]
    for i, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=i, value=h)
        cell.font = header_font
        cell.fill = header_fill

    txs = db.query(Transaction).filter(Transaction.year == year).order_by(
        Transaction.month, Transaction.category_id
    ).all()

    for row_idx, t in enumerate(txs, start=2):
        cat = cat_map.get(t.category_id)
        ws2.cell(row=row_idx, column=1, value=MESI[t.month - 1] if 1 <= t.month <= 12 else str(t.month))
        ws2.cell(row=row_idx, column=2, value=cat.name if cat else "?")
        ws2.cell(row=row_idx, column=3, value=cat.type if cat else "")
        ws2.cell(row=row_idx, column=4, value=t.description)
        ws2.cell(row=row_idx, column=5, value=t.amount)
        ws2.cell(row=row_idx, column=6, value=t.source or "manual")

    # Auto-width for sheet 2
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws2.column_dimensions[col].width = 18

    # Number format for currency
    for ws_sheet in [ws, ws2]:
        for row in ws_sheet.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, (int, float)) and cell.column > 1:
                    cell.number_format = '#,##0.00 €'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=bilancio_{year}.xlsx"},
    )


def export_year_csv(year: int, db: Session):
    """Fallback CSV export."""
    import csv

    txs = db.query(Transaction).filter(Transaction.year == year).order_by(
        Transaction.month, Transaction.category_id
    ).all()

    categories = {c.id: c for c in db.query(Category).all()}

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["Mese", "Categoria", "Tipo", "Descrizione", "Importo", "Fonte"])

    MESI = ["Gennaio", "Febbraio", "Marzo", "Aprile", "Maggio", "Giugno",
            "Luglio", "Agosto", "Settembre", "Ottobre", "Novembre", "Dicembre"]

    for t in txs:
        cat = categories.get(t.category_id)
        writer.writerow([
            MESI[t.month - 1] if 1 <= t.month <= 12 else str(t.month),
            cat.name if cat else "?",
            cat.type if cat else "",
            t.description,
            t.amount,
            t.source or "manual",
        ])

    content = buf.getvalue().encode("utf-8-sig")
    return StreamingResponse(
        io.BytesIO(content),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=bilancio_{year}.csv"},
    )
