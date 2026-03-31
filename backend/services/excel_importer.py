"""
Importa il file Excel Bilancino V8 nel database.
Struttura: foglio per anno, righe categoria, 2 colonne per mese (titolo+importo).
Col 0=Categoria, Col1=vuoto, Col2/3=Gen, Col4/5=Feb, ... Col24/25=Dic
"""
import re
from typing import Optional
import openpyxl
from sqlalchemy.orm import Session
from models import Category, Transaction, Investment

# Mappa 0-based: mese → (col_titolo, col_importo)
MONTH_COLS = {m: (2 + (m - 1) * 2, 3 + (m - 1) * 2) for m in range(1, 13)}

# Nomi riga Excel → nome categoria DB (uppercase match)
CATEGORY_ROW_MAP = {
    "GAS": "GAS",
    "LUCE": "LUCE",
    "ACQUA": "ACQUA",
    "VODAFONE": "VODAFONE",
    "NETFLIX": "NETFLIX",
    "SPESE ALIMENTARI": "SPESE ALIMENTARI",
    "AUTOMOBILE": "AUTOMOBILE",
    "SPESA SPORT": "SPESA SPORT",
    "USCITE E VACANZE": "USCITE E VACANZE",
    "TASSE": "TASSE",
    "TASSE ": "TASSE",
    "ALTRO": "ALTRO",
    "STIPENDIO": "STIPENDIO",
    "CONTRIBUTO MOGLIE": "CONTRIBUTO MOGLIE",
    "ALTRE ENTRATE": "ALTRE ENTRATE",
    "AFFITTO": "AFFITTO",
}

# Righe header/raggruppamento da ignorare (non sono dati)
SKIP_PREFIXES = {
    "BILANCIO", "CATEGORIA", "TITOLO SPESA",
    "SPESE FISSE", "ALTRE SPESE FISSE", "ALIMENTARI", "AUTO", "SPORT",
    "SVAGO/VACANZE", "SVAGO", "VACANZE",
    "ENTRATE / AFFITTO", "ENTRATE/AFFITTO",
    "INVESTIMENTI (INFO", "CHECK",
}

# Righe che segnalano fine dati
STOP_PREFIXES = {
    "TOTALE USCITE", "TOTALE ENTRATE", "RISPARMIO (ENTRATE",
    "RISPARMIO", "MESI",
}


def _normalize(val) -> str:
    if val is None:
        return ""
    return str(val).strip()


def _is_stop(v: str) -> bool:
    u = v.upper()
    return any(u.startswith(s) for s in STOP_PREFIXES)


def _is_skip(v: str) -> bool:
    u = v.upper()
    return any(u.startswith(s) for s in SKIP_PREFIXES)


def _detect_category(v: str) -> Optional[str]:
    """Ritorna il nome categoria DB se la cella è un header di categoria."""
    if not v:
        return None
    u = v.strip().upper()
    # Rimozione spazi extra
    u_clean = " ".join(u.split())
    return CATEGORY_ROW_MAP.get(u_clean)


def _get_cat_id(db: Session, name: str) -> Optional[int]:
    cat = db.query(Category).filter(Category.name == name).first()
    return cat.id if cat else None


def import_excel(filepath: str, db: Session) -> dict:
    wb = openpyxl.load_workbook(filepath, data_only=True)
    stats = []

    for sheet_name in wb.sheetnames:
        if not re.match(r"^\d{4}$", sheet_name):
            continue

        year = int(sheet_name)
        ws = wb[sheet_name]
        imported = 0
        skipped = 0
        current_category = None

        for row in ws.iter_rows(min_row=1, values_only=True):
            cell_a = _normalize(row[0])

            # Stop dell'import se arriviamo ai totali
            if _is_stop(cell_a):
                break

            # Salta righe header/raggruppamento
            if _is_skip(cell_a):
                continue

            # Controlla se è un nuovo header di categoria
            detected = _detect_category(cell_a)
            if detected:
                current_category = detected
                # Non fare break: la stessa riga può avere dati (es. VODAFONE col 1a spesa)

            if not current_category:
                continue

            cat_id = _get_cat_id(db, current_category)
            if not cat_id:
                skipped += 1
                continue

            # Leggi tutti i mesi
            for month, (ti, ai) in MONTH_COLS.items():
                if ai >= len(row):
                    continue

                raw_title = row[ti] if ti < len(row) else None
                raw_amount = row[ai]

                # Salta celle vuote
                if raw_amount is None:
                    continue

                try:
                    amount_val = float(raw_amount)
                except (ValueError, TypeError):
                    continue

                # Pulisci la descrizione: rimuovi prefisso "DD-" o "DD-" dal titolo
                title_str = _normalize(raw_title)
                # Rimuovi il prefisso giorno (es "18-Conad" → "Conad", "04-gpl" → "gpl")
                clean_title = re.sub(r"^\d{1,2}[-/]\s*", "", title_str).strip()
                if not clean_title:
                    clean_title = title_str
                description = clean_title if clean_title and clean_title.lower() != "none" \
                    else f"{current_category} {month}/{year}"

                # Evita duplicati esatti
                existing = db.query(Transaction).filter(
                    Transaction.year == year,
                    Transaction.month == month,
                    Transaction.category_id == cat_id,
                    Transaction.amount == amount_val,
                    Transaction.description == description,
                ).first()
                if existing:
                    skipped += 1
                    continue

                db.add(Transaction(
                    year=year,
                    month=month,
                    category_id=cat_id,
                    description=description,
                    amount=amount_val,
                    source="excel_import",
                ))
                imported += 1

        db.commit()
        stats.append({"year": year, "imported": imported, "skipped": skipped})

    inv_stats = _import_investments(wb, db)
    return {"years": stats, "investments": inv_stats}


def _import_investments(wb: openpyxl.Workbook, db: Session) -> dict:
    if "INVESTIMENTI" not in wb.sheetnames:
        return {"imported": 0}

    ws = wb["INVESTIMENTI"]
    imported = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 4:
            continue

        # Prova layout: asset=col1, data=col2, importo=col3
        asset = _normalize(row[1])
        date_val = row[2]
        amount = row[3]

        if not asset or not date_val or not amount:
            continue

        # Filtra righe intestazione
        if asset.upper() in ("ASSET", "PIATTAFORMA", "NOME", ""):
            continue

        try:
            amount_val = float(amount)
        except (ValueError, TypeError):
            continue

        # Tipo asset
        al = asset.lower()
        if any(k in al for k in ["etf", "scalable", "ishare"]):
            asset_type = "ETF"
        elif any(k in al for k in ["crypto", "bitcoin", "btc", "eth", "binance"]):
            asset_type = "Crypto"
        else:
            asset_type = "Altro"

        if hasattr(date_val, "strftime"):
            date_str = date_val.strftime("%Y-%m-%d")
        elif isinstance(date_val, str) and re.match(r"\d{2}/\d{2}/\d{4}", date_val):
            parts = date_val.split("/")
            date_str = f"{parts[2]}-{parts[1]}-{parts[0]}"
        else:
            continue

        existing = db.query(Investment).filter(
            Investment.date == date_str,
            Investment.asset == asset,
            Investment.amount_invested == abs(amount_val),
        ).first()
        if existing:
            continue

        db.add(Investment(
            date=date_str,
            asset=asset,
            asset_type=asset_type,
            amount_invested=abs(amount_val),
        ))
        imported += 1

    db.commit()
    return {"imported": imported}
